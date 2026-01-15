"""
ResultManager - Excel-Export fuer Walk-Forward Ergebnisse
"""

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any

import numpy as np
import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..core.logger import get_logger


@dataclass
class ExportConfig:
    """Konfiguration fuer Excel-Export."""

    # Allgemein
    output_dir: Path = field(default_factory=lambda: Path("results"))
    filename_prefix: str = "walkforward"
    include_timestamp: bool = True

    # Sheets aktivieren/deaktivieren
    include_summary: bool = True
    include_splits: bool = True
    include_trades: bool = True
    include_equity: bool = True
    include_config: bool = True
    include_predictions: bool = False  # Kann sehr gross werden

    # Formatierung
    use_formatting: bool = True
    chart_equity: bool = True

    def get_filename(self) -> str:
        """Generiert Dateinamen mit optionalem Timestamp."""
        if self.include_timestamp:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            return f"{self.filename_prefix}_{timestamp}.xlsx"
        return f"{self.filename_prefix}.xlsx"


class ResultManager:
    """
    Verwaltet Walk-Forward Ergebnisse und exportiert nach Excel.

    Features:
    - Multi-Sheet Excel mit Summary, Splits, Trades, Equity, Config
    - Formatierung mit Farben und Rahmen
    - Equity-Chart optional
    - Automatische Pfad-Generierung
    """

    # Farben fuer Formatierung
    COLORS = {
        'header': 'FF262626',      # Dunkel
        'header_font': 'FFFFFFFF', # Weiss
        'profit': 'FF33b34d',      # Gruen
        'loss': 'FFcc4d33',        # Rot
        'neutral': 'FF808080',     # Grau
        'warning': 'FFe6b333',     # Orange
        'alternating': 'FFF5F5F5', # Hellgrau
    }

    def __init__(self, config: Optional[ExportConfig] = None):
        """
        Initialisiert den ResultManager.

        Args:
            config: Export-Konfiguration (optional)
        """
        self.config = config or ExportConfig()
        self.logger = get_logger()

        if not OPENPYXL_AVAILABLE:
            self.logger.warning("openpyxl nicht installiert - Excel-Export deaktiviert")

    def export_results(
        self,
        result: 'WalkForwardResult',
        config: Optional['WalkForwardConfig'] = None,
        output_path: Optional[Path] = None
    ) -> Optional[Path]:
        """
        Exportiert Walk-Forward Ergebnisse nach Excel.

        Args:
            result: WalkForwardResult mit allen Daten
            config: WalkForwardConfig fuer Config-Sheet
            output_path: Optionaler Pfad (ueberschreibt config)

        Returns:
            Pfad zur erstellten Datei oder None bei Fehler
        """
        if not OPENPYXL_AVAILABLE:
            self.logger.error("Excel-Export nicht moeglich: openpyxl fehlt")
            return None

        try:
            # Pfad bestimmen
            if output_path is None:
                self.config.output_dir.mkdir(parents=True, exist_ok=True)
                output_path = self.config.output_dir / self.config.get_filename()

            self.logger.info(f"Exportiere Ergebnisse nach: {output_path}")

            # Workbook erstellen
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Default-Sheet entfernen

            # Sheets erstellen
            if self.config.include_summary:
                self._create_summary_sheet(wb, result)

            if self.config.include_splits and result.split_results:
                self._create_splits_sheet(wb, result)

            if self.config.include_trades:
                self._create_trades_sheet(wb, result)

            if self.config.include_equity:
                self._create_equity_sheet(wb, result)

            if self.config.include_config and config:
                self._create_config_sheet(wb, config)

            if self.config.include_predictions:
                self._create_predictions_sheet(wb, result)

            # Speichern
            wb.save(output_path)
            self.logger.success(f"Excel-Export erfolgreich: {output_path}")

            return output_path

        except Exception as e:
            self.logger.error(f"Excel-Export fehlgeschlagen: {e}")
            return None

    def _create_summary_sheet(self, wb: 'openpyxl.Workbook', result: 'WalkForwardResult'):
        """Erstellt Summary-Sheet mit Gesamtstatistiken."""
        ws = wb.create_sheet("Summary")

        # Header
        ws['A1'] = "Walk-Forward Analyse - Zusammenfassung"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        # Metriken
        metrics = [
            ("", ""),  # Leerzeile
            ("Zeitraum", ""),
            ("Start", result.start_date.strftime("%Y-%m-%d %H:%M") if result.start_date else "N/A"),
            ("Ende", result.end_date.strftime("%Y-%m-%d %H:%M") if result.end_date else "N/A"),
            ("Anzahl Splits", str(result.total_splits)),
            ("", ""),
            ("Performance", ""),
            ("Total Return", f"{result.total_return:.2%}"),
            ("Sharpe Ratio", f"{result.sharpe_ratio:.3f}"),
            ("Max Drawdown", f"{result.max_drawdown:.2%}"),
            ("Win Rate", f"{result.win_rate:.2%}"),
            ("Profit Factor", f"{result.profit_factor:.2f}"),
            ("", ""),
            ("Trades", ""),
            ("Anzahl Trades", str(result.total_trades)),
            ("Gewinn-Trades", str(result.winning_trades)),
            ("Verlust-Trades", str(result.losing_trades)),
            ("Durchschn. Trade", f"{result.avg_trade:.4f}"),
            ("", ""),
            ("Laufzeit", ""),
            ("Dauer", f"{result.execution_time:.1f} Sekunden"),
        ]

        for i, (label, value) in enumerate(metrics, start=3):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value

            if label and not value:  # Kategorie-Header
                ws[f'A{i}'].font = Font(bold=True)

        # Spaltenbreiten
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25

        self._apply_formatting(ws) if self.config.use_formatting else None

    def _create_splits_sheet(self, wb: 'openpyxl.Workbook', result: 'WalkForwardResult'):
        """Erstellt Sheet mit Split-Details."""
        ws = wb.create_sheet("Splits")

        # Header
        headers = [
            "Split", "Train Start", "Train Ende", "Test Start", "Test Ende",
            "Samples", "Return", "Sharpe", "Max DD", "Trades", "Win Rate"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            if self.config.use_formatting:
                cell.font = Font(bold=True, color=self.COLORS['header_font'])
                cell.fill = PatternFill(start_color=self.COLORS['header'],
                                       end_color=self.COLORS['header'],
                                       fill_type='solid')

        # Daten
        for i, split in enumerate(result.split_results, start=2):
            ws.cell(row=i, column=1, value=split.split_id)
            ws.cell(row=i, column=2, value=split.train_start.strftime("%Y-%m-%d") if split.train_start else "")
            ws.cell(row=i, column=3, value=split.train_end.strftime("%Y-%m-%d") if split.train_end else "")
            ws.cell(row=i, column=4, value=split.test_start.strftime("%Y-%m-%d") if split.test_start else "")
            ws.cell(row=i, column=5, value=split.test_end.strftime("%Y-%m-%d") if split.test_end else "")
            ws.cell(row=i, column=6, value=split.n_test_samples)

            # Return mit Farbcodierung
            return_cell = ws.cell(row=i, column=7, value=f"{split.total_return:.2%}")
            if self.config.use_formatting:
                if split.total_return > 0:
                    return_cell.font = Font(color=self.COLORS['profit'][2:])
                elif split.total_return < 0:
                    return_cell.font = Font(color=self.COLORS['loss'][2:])

            ws.cell(row=i, column=8, value=f"{split.sharpe_ratio:.3f}")
            ws.cell(row=i, column=9, value=f"{split.max_drawdown:.2%}")
            ws.cell(row=i, column=10, value=split.n_trades)
            ws.cell(row=i, column=11, value=f"{split.win_rate:.1%}")

            # Alternating row colors
            if self.config.use_formatting and i % 2 == 0:
                for col in range(1, 12):
                    ws.cell(row=i, column=col).fill = PatternFill(
                        start_color=self.COLORS['alternating'],
                        end_color=self.COLORS['alternating'],
                        fill_type='solid'
                    )

        # Auto-Spaltenbreite
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

    def _create_trades_sheet(self, wb: 'openpyxl.Workbook', result: 'WalkForwardResult'):
        """Erstellt Sheet mit allen Trades."""
        ws = wb.create_sheet("Trades")

        # Header
        headers = [
            "Split", "Trade #", "Einstieg", "Ausstieg", "Richtung",
            "Entry Price", "Exit Price", "PnL", "PnL %", "Bars"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            if self.config.use_formatting:
                cell.font = Font(bold=True, color=self.COLORS['header_font'])
                cell.fill = PatternFill(start_color=self.COLORS['header'],
                                       end_color=self.COLORS['header'],
                                       fill_type='solid')

        # Alle Trades sammeln
        row = 2
        for split in result.split_results:
            for trade in split.trades:
                ws.cell(row=row, column=1, value=split.split_id)
                ws.cell(row=row, column=2, value=trade.trade_id)
                ws.cell(row=row, column=3, value=trade.entry_time.strftime("%Y-%m-%d %H:%M") if trade.entry_time else "")
                ws.cell(row=row, column=4, value=trade.exit_time.strftime("%Y-%m-%d %H:%M") if trade.exit_time else "")
                ws.cell(row=row, column=5, value=trade.direction)
                ws.cell(row=row, column=6, value=f"{trade.entry_price:.2f}")
                ws.cell(row=row, column=7, value=f"{trade.exit_price:.2f}")

                # PnL mit Farbcodierung
                pnl_cell = ws.cell(row=row, column=8, value=f"{trade.pnl:.2f}")
                pnl_pct_cell = ws.cell(row=row, column=9, value=f"{trade.pnl_percent:.2%}")

                if self.config.use_formatting:
                    color = self.COLORS['profit'][2:] if trade.pnl > 0 else self.COLORS['loss'][2:]
                    pnl_cell.font = Font(color=color)
                    pnl_pct_cell.font = Font(color=color)

                ws.cell(row=row, column=10, value=trade.bars_held)
                row += 1

        # Auto-Spaltenbreite
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

    def _create_equity_sheet(self, wb: 'openpyxl.Workbook', result: 'WalkForwardResult'):
        """Erstellt Equity-Sheet mit Chart."""
        ws = wb.create_sheet("Equity")

        # Header
        headers = ["Timestamp", "Equity", "Drawdown", "Split"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            if self.config.use_formatting:
                cell.font = Font(bold=True, color=self.COLORS['header_font'])
                cell.fill = PatternFill(start_color=self.COLORS['header'],
                                       end_color=self.COLORS['header'],
                                       fill_type='solid')

        # Equity-Daten sammeln
        row = 2
        for split in result.split_results:
            for point in split.equity_curve:
                ws.cell(row=row, column=1, value=point.timestamp.strftime("%Y-%m-%d %H:%M") if point.timestamp else "")
                ws.cell(row=row, column=2, value=point.equity)
                ws.cell(row=row, column=3, value=point.drawdown)
                ws.cell(row=row, column=4, value=split.split_id)
                row += 1

        # Chart erstellen wenn aktiviert
        if self.config.chart_equity and row > 2:
            chart = LineChart()
            chart.title = "Equity Curve"
            chart.style = 10
            chart.y_axis.title = "Equity"
            chart.x_axis.title = "Time"
            chart.width = 20
            chart.height = 10

            # Daten-Referenz
            data = Reference(ws, min_col=2, min_row=1, max_row=row-1)
            chart.add_data(data, titles_from_data=True)

            ws.add_chart(chart, "F2")

        # Spaltenbreiten
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10

    def _create_config_sheet(self, wb: 'openpyxl.Workbook', config: 'WalkForwardConfig'):
        """Erstellt Config-Sheet mit allen Parametern."""
        ws = wb.create_sheet("Config")

        # Header
        ws['A1'] = "Parameter"
        ws['B1'] = "Wert"

        if self.config.use_formatting:
            for cell in [ws['A1'], ws['B1']]:
                cell.font = Font(bold=True, color=self.COLORS['header_font'])
                cell.fill = PatternFill(start_color=self.COLORS['header'],
                                       end_color=self.COLORS['header'],
                                       fill_type='solid')

        # Config-Werte aus Dataclass extrahieren
        row = 2
        config_dict = config.__dict__ if hasattr(config, '__dict__') else {}

        for key, value in config_dict.items():
            ws.cell(row=row, column=1, value=key)

            # Wert formatieren
            if isinstance(value, (list, tuple)):
                ws.cell(row=row, column=2, value=str(value))
            elif isinstance(value, bool):
                ws.cell(row=row, column=2, value="Ja" if value else "Nein")
            elif isinstance(value, float):
                ws.cell(row=row, column=2, value=f"{value:.4f}")
            elif hasattr(value, 'value'):  # Enum
                ws.cell(row=row, column=2, value=value.value)
            elif hasattr(value, '__name__'):  # Typ
                ws.cell(row=row, column=2, value=value.__name__)
            else:
                ws.cell(row=row, column=2, value=str(value))

            row += 1

        # Spaltenbreiten
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 40

    def _create_predictions_sheet(self, wb: 'openpyxl.Workbook', result: 'WalkForwardResult'):
        """Erstellt Sheet mit allen Predictions (kann sehr gross sein)."""
        ws = wb.create_sheet("Predictions")

        # Header
        headers = ["Split", "Bar", "Timestamp", "Pred_Hold", "Pred_Buy", "Pred_Sell", "Signal", "True Label"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            if self.config.use_formatting:
                cell.font = Font(bold=True, color=self.COLORS['header_font'])
                cell.fill = PatternFill(start_color=self.COLORS['header'],
                                       end_color=self.COLORS['header'],
                                       fill_type='solid')

        # Predictions sammeln
        row = 2
        for split in result.split_results:
            if hasattr(split, 'predictions') and split.predictions is not None:
                for i, pred in enumerate(split.predictions):
                    ws.cell(row=row, column=1, value=split.split_id)
                    ws.cell(row=row, column=2, value=i)
                    # Weitere Spalten je nach Datenstruktur
                    row += 1

    def _apply_formatting(self, ws):
        """Wendet Standard-Formatierung auf ein Sheet an."""
        # Rahmen fuer alle Zellen mit Daten
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border

    def export_to_csv(
        self,
        result: 'WalkForwardResult',
        output_dir: Optional[Path] = None
    ) -> Dict[str, Path]:
        """
        Exportiert Ergebnisse als separate CSV-Dateien.

        Args:
            result: WalkForwardResult
            output_dir: Ausgabe-Verzeichnis

        Returns:
            Dict mit Sheet-Namen und Pfaden
        """
        output_dir = output_dir or self.config.output_dir
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        exported = {}

        # Summary
        summary_data = {
            'Metric': ['Total Return', 'Sharpe Ratio', 'Max Drawdown', 'Win Rate',
                      'Profit Factor', 'Total Trades', 'Winning Trades', 'Losing Trades'],
            'Value': [result.total_return, result.sharpe_ratio, result.max_drawdown,
                     result.win_rate, result.profit_factor, result.total_trades,
                     result.winning_trades, result.losing_trades]
        }
        summary_path = output_dir / f"summary_{timestamp}.csv"
        pd.DataFrame(summary_data).to_csv(summary_path, index=False)
        exported['summary'] = summary_path

        # Splits
        if result.split_results:
            splits_data = []
            for split in result.split_results:
                splits_data.append({
                    'split_id': split.split_id,
                    'train_start': split.train_start,
                    'train_end': split.train_end,
                    'test_start': split.test_start,
                    'test_end': split.test_end,
                    'n_test_samples': split.n_test_samples,
                    'total_return': split.total_return,
                    'sharpe_ratio': split.sharpe_ratio,
                    'max_drawdown': split.max_drawdown,
                    'n_trades': split.n_trades,
                    'win_rate': split.win_rate
                })
            splits_path = output_dir / f"splits_{timestamp}.csv"
            pd.DataFrame(splits_data).to_csv(splits_path, index=False)
            exported['splits'] = splits_path

        # Trades
        trades_data = []
        for split in result.split_results:
            for trade in split.trades:
                trades_data.append({
                    'split_id': split.split_id,
                    'trade_id': trade.trade_id,
                    'entry_time': trade.entry_time,
                    'exit_time': trade.exit_time,
                    'direction': trade.direction,
                    'entry_price': trade.entry_price,
                    'exit_price': trade.exit_price,
                    'pnl': trade.pnl,
                    'pnl_percent': trade.pnl_percent,
                    'bars_held': trade.bars_held
                })

        if trades_data:
            trades_path = output_dir / f"trades_{timestamp}.csv"
            pd.DataFrame(trades_data).to_csv(trades_path, index=False)
            exported['trades'] = trades_path

        # Equity
        equity_data = []
        for split in result.split_results:
            for point in split.equity_curve:
                equity_data.append({
                    'split_id': split.split_id,
                    'timestamp': point.timestamp,
                    'equity': point.equity,
                    'drawdown': point.drawdown
                })

        if equity_data:
            equity_path = output_dir / f"equity_{timestamp}.csv"
            pd.DataFrame(equity_data).to_csv(equity_path, index=False)
            exported['equity'] = equity_path

        self.logger.success(f"CSV-Export abgeschlossen: {len(exported)} Dateien")
        return exported
